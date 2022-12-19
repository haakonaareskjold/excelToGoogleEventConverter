import openpyxl
import os.path
import tkinter as tk
from datetime import datetime
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from tkinter import filedialog


def main():
    # Set up the tkinter file dialog
    root = tk.Tk()
    root.withdraw()

    # Open the file dialog and get the file path
    file_path = filedialog.askopenfilename()

    # Read the excel file and store the data in a list
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet1']  # Read the excel file and store the data in a list
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Sheet1']

    # Get the dates and names from the excel sheet
    names = []
    releaseDates = []
    preorderDates = []

    for row in sheet.iter_rows(min_row=2):
        name = row[0].value
        releaseDate = row[1].value
        preorderDate = row[2].value
        names.append(name)

        if releaseDate is not None:
            releaseDates.append(releaseDate)
        else:
            releaseDates.append(datetime.fromtimestamp(3600))

        if preorderDate is not None:
            preorderDates.append(preorderDate)
        else:
            preorderDates.append(datetime.fromtimestamp(3600))

    # Convert the dates to a datetime object
    releaseDates = [
        datetime.strftime(
            releaseDate,
            '%Y-%m-%d') for releaseDate in releaseDates]
    preorderDates = [
        datetime.strftime(
            preorderDate,
            '%Y-%m-%d') for preorderDate in preorderDates]

    # If modifying these scopes, delete the file token.json.
    SCOPES = ['https://www.googleapis.com/auth/calendar']

    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            # If modifying these scopes, delete the file token.json.
            token.write(creds.to_json())

    # Set up the Google Calendar API
    service = build(
        'calendar',
        'v3',
        credentials=creds,
        static_discovery=False)

    # Create the events in the calendar
    calendar_id = 'primary'
    for name, releaseDate, preorderDate in zip(
            names, releaseDates, preorderDates):
        mangaRelease = {
            'summary': name + ' release',
            'start': {
                'date': releaseDate
            },
            'end': {
                'date': releaseDate
            },
        }
        mangaPreorder = {
            'summary': name + ' preorder',
            'start': {
                'date': preorderDate
            },
            'end': {
                'date': preorderDate
            },
        }

        # get a list of events with freesearch matching name of mangas
        results = service.events().list(calendarId=calendar_id, q=name).execute()
        events = results.get('items', [])

        existingMangas = None
        for event in events:
            existingMangas = event['summary']

        if existingMangas is None:
            # Create the events if release or preorder is in future
            if mangaRelease['start']['date'] >= datetime.now().strftime(
                    "%Y-%m-%d"):
                service.events().insert(
                    calendarId=calendar_id, body=mangaRelease).execute()
                print(
                    'Event created for release: ' +
                    mangaRelease['summary'] +
                    ' on date ' +
                    releaseDate)
            if mangaPreorder['start']['date'] >= datetime.now().strftime(
                    "%Y-%m-%d"):
                service.events().insert(
                    calendarId=calendar_id, body=mangaPreorder).execute()
                print(
                    'Event created for preorder: ' +
                    mangaPreorder['summary'] +
                    ' on date ' +
                    preorderDate)
        else:
            print('event already exists:' + mangaRelease['summary'])


if __name__ == "__main__":
    main()
